from fastapi import APIRouter, HTTPException, Depends
from pydantic import BaseModel
from typing import Dict, Any
from sqlmodel import Session
from ..settings_manager import settings_manager
from ..routers.auth import get_current_user
from ..db import get_session
from ..config import get_settings as get_env_settings
from ..lib.network_file_saver import NetworkFileSaver

router = APIRouter(
    prefix="/settings",
    tags=["settings"],
    responses={404: {"description": "Not found"}},
)

class LoginRequest(BaseModel):
    username: str
    password: str

@router.post("/login", response_model=Dict[str, bool])
async def login_settings(credentials: LoginRequest):
    """
    Verify settings page credentials.
    """
    env_settings = get_env_settings()

    # Get credentials from settings (loaded from env)
    valid_user = env_settings.USER_SETARI
    valid_pass = env_settings.PASS_SETARI

    if not valid_user or not valid_pass:
        # Fail safe: if not configured in .env, deny access
        raise HTTPException(status_code=500, detail="Server misconfiguration: Credentials not set")

    if credentials.username == valid_user and credentials.password == valid_pass:
        return {"success": True}

    raise HTTPException(status_code=401, detail="Invalid credentials")


@router.get("/", response_model=Dict[str, Any])
async def get_settings():
    """
    Get all current settings.
    """
    return settings_manager.get_settings()

@router.put("/", response_model=Dict[str, Any])
async def update_settings(new_settings: Dict[str, Any]):
    """
    Update settings.
    """
    try:
        from ..settings_manager import logger as settings_logger
        settings_logger.info(f"Received update_settings request. Payload size: {len(str(new_settings))} chars")
        settings_manager.save_settings(new_settings)
        return settings_manager.get_settings()
    except Exception as e:
        settings_logger.error(f"Exception in update_settings endpoint: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/reset", response_model=Dict[str, Any])
async def reset_settings():
    """
    Reset settings to defaults.
    """
    try:
        settings_manager.reset_to_defaults()
        return settings_manager.get_settings()
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/precalculate-models-codes", response_model=Dict[str, Any])
async def precalculate_models_codes(
    restart: bool = False
):
    """
    Triggers pre-calculation of models and codes for all cases.
    This is a resource-intensive operation that should be run during off-peak hours.
    Runs in a background thread to avoid blocking the API.

    Query Parameters:
        restart: If True, reset all precalculated data and start from scratch.
                If False, only process incomplete cases (resume mode).
    """
    from ..db import engine as main_engine
    from ..db_modele import modele_engine
    from ..db_coduri import coduri_engine
    from sqlmodel import Session
    from ..logic.precalculation_service import precalculate_models_and_codes, precalc_status, precalc_status_lock
    import logging
    import threading

    logger = logging.getLogger(__name__)
    logger.info(f"Pre-calculation endpoint called, restart={restart}")

    # Check if already running
    with precalc_status_lock:
        if precalc_status['is_running']:
            return {
                'success': False,
                'message': 'A precalculation process is already running',
                'is_running': True
            }

    def run_precalculation():
        """Background thread function to run precalculation"""
        try:
            with Session(main_engine) as main_session, \
                 Session(modele_engine) as modele_session, \
                 Session(coduri_engine) as coduri_session:

                results = precalculate_models_and_codes(
                    main_session=main_session,
                    modele_session=modele_session,
                    coduri_session=coduri_session,
                    batch_size=100,
                    limit_modele=5,
                    limit_coduri=5,
                    restart_from_zero=restart
                )
                logger.info(f"Precalculation completed: {results}")
        except Exception as e:
            logger.error(f"Error in background precalculation: {e}", exc_info=True)

    # Start the background thread
    thread = threading.Thread(target=run_precalculation, daemon=True)
    thread.start()

    return {
        'success': True,
        'message': 'Precalculation started in background',
        'is_running': True,
        'restart_mode': restart
    }


@router.get("/precalculate-status", response_model=Dict[str, Any])
async def get_precalculate_status():
    """
    Get the current status of the precalculation process.
    """
    from ..db import engine as main_engine
    from sqlmodel import Session
    from ..logic.precalculation_service import get_precalculation_status

    try:
        with Session(main_engine) as main_session:
            status = get_precalculation_status(main_session)
            return status
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Error getting precalculation status: {str(e)}"
        )


@router.post("/precalculate-stop", response_model=Dict[str, Any])
async def stop_precalculate():
    """
    Stop the currently running precalculation process.
    """
    from ..logic.precalculation_service import stop_precalculation
    import logging

    logger = logging.getLogger(__name__)
    logger.info("Stop precalculation requested")

    try:
        result = stop_precalculation()
        return result
    except Exception as e:
        logger.error(f"Error stopping precalculation: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Error stopping precalculation: {str(e)}"
        )


@router.get("/export-llm-data", response_model=Dict[str, Any])
async def export_llm_data(
    session: Session = Depends(get_session)
):
    """
    Export fact situations from last search query for LLM refinement.
    Returns JSON with case IDs, names, and full fact situations.
    """
    from ..models import UltimaInterogare
    import logging

    logger = logging.getLogger(__name__)

    try:
        # Get last query data
        ultima = session.get(UltimaInterogare, 1)

        if not ultima or not ultima.speta_ids:
            return {
                'success': False,
                'message': 'Nu există rezultate salvate din ultima căutare.',
                'query_text': '',
                'total_spete': 0,
                'spete': []
            }

        # Use helper to generate data and prompt
        spete_export, optimized_prompt = _generate_llm_data(session, ultima)

        logger.info(f"Exported {len(spete_export)} cases for LLM from last query: '{ultima.query_text[:50]}'")

        return {
            'success': True,
            'query_text': ultima.query_text,
            'total_spete': len(spete_export),
            'spete': spete_export,
            'optimized_prompt': optimized_prompt
        }

    except Exception as e:
        logger.error(f"Error exporting LLM data: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Eroare la exportul datelor: {str(e)}"
        )


@router.post("/analyze-llm-data", response_model=Dict[str, Any])
async def analyze_llm_data(
    session: Session = Depends(get_session)
):
    """
    Start LLM analysis and return job_id immediately.
    Client should poll /analyze-llm-status/{job_id} for results.
    """
    from ..models import UltimaInterogare
    from ..logic.queue_manager import queue_manager
    import logging
    import httpx

    logger = logging.getLogger(__name__)

    try:
        # Get last query data
        ultima = session.get(UltimaInterogare, 1)

        if not ultima or not ultima.speta_ids:
            return {
                'success': False,
                'message': 'Nu există rezultate salvate din ultima căutare.',
            }

        # Check network settings first to determine candidate count and mode
        network_enabled = settings_manager.get_value('setari_retea', 'retea_enabled', False)

        # Check if Pro Search was used in the original query
        is_pro_search = getattr(ultima, 'pro_search', False)

        custom_template = None
        if network_enabled and is_pro_search:
            # SPECIAL MODE: Pro Search + Network Mode
            # Use 50 candidates for maximum coverage
            candidate_count = 50
            logger.info(f"[AI FILTERING] PRO + NETWORK MODE ACTIVATED: Using {candidate_count} candidates")

            # Strict prompt for Pro mode: max 10 relevant results, strict formatting
            custom_template = (
                "Esti un judecator cu experienta, capabil sa analizeze spete juridice complexe si sa identifice precedente relevante.\\n\\n"
                "SARCINA TA:\\n"
                "Analizeaza situatia de fapt prezentata de justitiabil mai jos si compar-o cu cele {num_candidates} spete candidate furnizate.\\n\\n"
                "Selecteaza DOAR spetele strict relevante pentru situatia justitiabilului (maxim 10 spete).\\n"
                "- Daca gasesti 10 sau mai multe spete relevante, returneaza EXACT 10 spete\\n"
                "- Daca gasesti mai putin de 10 spete relevante, returneaza DOAR cele relevante (ex: 3, 5, 7 spete este acceptabil)\\n"
                "- Ordoneaza rezultatele de la CEL MAI RELEVANT la CEL MAI PUTIN RELEVANT\\n"
                "- Evalueaza relevanta luand in considerare: situatia de fapt, obiectul si elementele de individualizare\\n\\n"
                "SITUATIA DE FAPT A JUSTITIABILULUI:\\n"
                "\"{query_text}\"\\n\\n"
                "LISTA DE SPETE CANDIDATE ({num_candidates} Spete Pre-filtrate):\\n"
                "{prompt_spete}\\n\\n"
                "FORMATUL RASPUNSULUI - FOARTE IMPORTANT:\\n"
                "Genereaza EXCLUSIV ID-urile spetelor selectate, separate prin virgula.\\n"
                "NU adauga NICIUN alt text, comentariu, explicatie sau introducere.\\n"
                "NU adauga text inainte, la mijloc sau dupa numerele de ID.\\n"
                "Raspunsul trebuie sa contina DOAR numere separate prin virgule.\\n\\n"
                "Exemple de raspunsuri CORECTE:\\n"
                "- Pentru 1 rezultat: 123\\n"
                "- Pentru 3 rezultate: 123, 456, 789\\n"
                "- Pentru 10 rezultate: 123, 456, 789, 1011, 1213, 1415, 1617, 1819, 2021, 2223\\n\\n"
                "Exemple de raspunsuri INCORECTE (NU face asta):\\n"
                "- \"Iata spetele relevante: 123, 456\" (are text inainte)\\n"
                "- \"123, 456 - acestea sunt cele mai relevante\" (are text dupa)\\n"
                "- \"Am gasit urmatoarele 3 spete: 123, 456, 789\" (are text inainte)"
            )
        elif network_enabled:
            # NORMAL NETWORK MODE (without Pro Search)
            # Use more candidates for network mode (default 50 or max available)
            candidate_count = settings_manager.get_value('setari_generale', 'top_k_results', 50)
            logger.info(f"[AI FILTERING] NETWORK MODE: Using {candidate_count} candidates")

            # Original network mode prompt
            custom_template = (
                "Esti un judecator cu experienta, capabil sa analizeze spete juridice complexe si sa identifice precedente relevante.\\n\\n"
                "SARCINA TA:\\n"
                "Analizeaza situatia de fapt prezentata de justitiabil mai jos si compar-o cu cele {num_candidates} spete candidate furnizate.\\n"
                "Selecteaza cele mai relevante 10 spete asemanatoare cu situatia de fapt a utilizatorului, luand in considerare situatia de fapt, obiectul si elementele de individualizare.\\n"
                "Daca nu gasesti 10 spete strict relevante (ex: alta materie sau alt obiect), afiseaza TOATE spetele pe care le consideri relevante, "
                "ordonate de la cea mai relevanta la cea mai putin relevanta.\\n\\n"
                "SITUATIA DE FAPT A JUSTITIABILULUI:\\n"
                "\"{query_text}\"\\n\\n"
                "LISTA DE SPETE CANDIDATE ({num_candidates} Spete Pre-filtrate):\\n"
                "{prompt_spete}\\n\\n"
                "FORMATUL RASPUNSULUI:\\n"
                "Genereaza EXCLUSIV ID-urile spetelor selectate, separate prin virgula.\\n"
                "Nu adauga niciun alt text, comentariu, explicatie sau introducere - nici inainte, nici dupa.\\n"
                "Exemplu de raspuns valid pentru 1 rezultat: 123\\n"
                "Exemplu de raspuns valid pentru 3 rezultate: 123, 456, 789\\n"
                "Exemplu de raspuns valid pentru 10 rezultate: 123, 456, 789, 1011, 1213, 1415, 1617, 1819, 2021, 2223"
            )
        else:
            # NORMAL MODE (no network)
            candidate_count = settings_manager.get_value('setari_llm', 'ai_filtering_llm_candidate_count', 5)
            logger.info(f"[AI FILTERING] NORMAL MODE: Using {candidate_count} candidates")

        # Generate the prompt using the helper
        all_candidates, optimized_prompt = _generate_llm_data(
            session,
            ultima,
            candidate_count=candidate_count,
            custom_template=custom_template
        )

        # Define async processor function
        async def process_llm_analysis(payload: dict):
            """Process the LLM analysis."""

            # ===== SALVARE PROMPT ÎN REȚEA ÎNAINTE DE LLM =====
            logger.info("[AI FILTERING] Verificăm configurarea salvării în rețea...")
            network_enabled = settings_manager.get_value('setari_retea', 'retea_enabled', False)

            if network_enabled:
                logger.info("[AI FILTERING] ✓ Salvarea în rețea este ACTIVATĂ")

                # Preluăm setările de rețea
                retea_host = settings_manager.get_value('setari_retea', 'retea_host', '')
                retea_folder = settings_manager.get_value('setari_retea', 'retea_folder_partajat', '')
                retea_subfolder = settings_manager.get_value('setari_retea', 'retea_subfolder', '')

                logger.info(f"[AI FILTERING] Configurare rețea: \\\\{retea_host}\\{retea_folder}")

                # Validăm configurarea
                # Permitem host gol dacă folderul este o cale absolută (pentru mount-uri locale)
                import os
                is_local_path = os.path.isabs(retea_folder) if retea_folder else False

                if (not retea_host and not is_local_path) or not retea_folder:
                    error_msg = "Salvarea în rețea este activată, dar configurația este incompletă (lipsește host sau folder partajat)."
                    logger.error(f"[AI FILTERING] ❌ EROARE CONFIGURARE: {error_msg}")
                    # Nu ridicăm excepție aici pentru a nu bloca worker-ul, dar returnăm eroare
                    return {
                        'success': False,
                        'error': error_msg
                    }

                # Salvăm promptul în rețea
                logger.info("[AI FILTERING] Începem salvarea promptului în rețea...")
                success, message, saved_path = NetworkFileSaver.save_to_network(
                    content=payload['prompt'],
                    host=retea_host,
                    shared_folder=retea_folder,
                    subfolder=retea_subfolder
                )

                if not success:
                    logger.error(f"[AI FILTERING] ❌ EROARE SALVARE REȚEA: {message}")
                    return {
                        'success': False,
                        'error': f"Eroare la salvarea în rețea: {message}"
                    }

                logger.info(f"[AI FILTERING] ✅ Prompt salvat cu succes: {saved_path}")

                # ===== POLLING PENTRU RĂSPUNS =====
                logger.info("[AI FILTERING] Începem polling pentru răspuns din rețea...")

                # Polling pentru fișierul de răspuns
                poll_success, poll_content, response_path = await NetworkFileSaver.poll_for_response(
                    saved_path=saved_path,
                    timeout_seconds=1200,  # 20 minute
                    poll_interval=10  # 10 secunde
                )

                if not poll_success:
                    # Timeout sau altă eroare
                    logger.error(f"[AI FILTERING] ❌ POLLING FAILED: {poll_content}")
                    return {
                        'success': False,
                        'error': poll_content
                    }

                logger.info(f"[AI FILTERING] ✅ Răspuns primit din rețea!")

                # ===== PARSARE ID-URI SPEȚE DIN RĂSPUNS =====
                logger.info("[AI FILTERING] Parsăm ID-urile spețelor din răspuns...")

                import re
                id_matches = re.findall(r'\d+', poll_content)
                ai_selected_ids = [int(id_str) for id_str in id_matches]

                logger.info(f"[AI FILTERING] ✓ ID-uri extrase: {ai_selected_ids}")

                # ===== ȘTERGERE FIȘIER RĂSPUNS =====
                logger.info("[AI FILTERING] Curățăm fișierul de răspuns...")

                delete_success, delete_message = NetworkFileSaver.delete_response_file(response_path)

                if delete_success:
                    logger.info(f"[AI FILTERING] ✓ {delete_message}")
                else:
                    logger.warning(f"[AI FILTERING] ⚠️ Eroare la ștergere (non-critică): {delete_message}")

                # ===== RETURNARE REZULTATE =====
                logger.info("[AI FILTERING] ✅ Procesare completă - returnăm rezultatele")
                return {
                    'success': True,
                    'response': poll_content,
                    'ai_selected_ids': ai_selected_ids,
                    'all_candidates': payload.get('all_candidates', []),
                    'network_save': True,
                    'saved_path': saved_path
                }

            # ===== CONTINUARE CU TRIMITEREA CĂTRE LLM (DOAR DACĂ REȚEA E OFF) =====
            logger.info("[AI FILTERING] Salvarea în rețea este DEZACTIVATĂ, se continuă cu LLM local")

            # Get LLM URL from settings
            llm_url = settings_manager.get_value('setari_llm', 'llm_url')

            # If not set (or empty string from default), use env or safe fallback
            if not llm_url:
                env_settings = get_env_settings()
                if env_settings.OLLAMA_URL:
                    llm_url = f"{env_settings.OLLAMA_URL}/api/generate"
                else:
                    llm_url = 'http://192.168.1.30:11434/api/generate'

            logger.info(f"Sending request to LLM at {llm_url}...")

            # --- CONFIGURARE MODEL QWEN ---
            llm_payload = {
                "model": "verdict-ro",  # Folosim modelul Qwen instalat
                "prompt": payload['prompt'],
                "stream": False,                 # False pentru raspuns complet
                "max_tokens": 2048,              # Limita marita pentru raspunsuri complexe
                "temperature": 0.1               # Temperatura mica pentru precizie
            }
            # ------------------------------

            async with httpx.AsyncClient(timeout=1200.0) as client:
                response = await client.post(llm_url, json=llm_payload)
                response.raise_for_status()
                result = response.json()

            logger.info("Received response from LLM")

            # Parse AI-selected IDs from response (can be single or comma-separated)
            llm_response_text = result.get('response', '')
            ai_selected_ids = []

            # Extract all numbers from response
            import re
            id_matches = re.findall(r'\d+', llm_response_text)
            ai_selected_ids = [int(id_str) for id_str in id_matches]

            logger.info(f"LLM selected IDs: {ai_selected_ids}")

            return {
                'success': True,
                'response': llm_response_text,
                'ai_selected_ids': ai_selected_ids,
                'all_candidates': payload.get('all_candidates', []),
                'full_response': result
            }

        # Add to queue and get job_id immediately
        payload = {
            'prompt': optimized_prompt,
            'all_candidates': all_candidates  # Include all candidates for response
        }
        job_id, _ = await queue_manager.add_to_queue(payload, process_llm_analysis)

        logger.info(f"LLM analysis queued with job_id: {job_id}")

        return {
            'success': True,
            'job_id': job_id,
            'message': 'Analiză pusă în coadă. Folosește job_id pentru a verifica statusul.'
        }

    except RuntimeError as e:
        logger.error(f"Queue error: {e}")
        raise HTTPException(status_code=503, detail=f"Coada este plină. Vă rugăm să încercați din nou mai târziu.")
    except Exception as e:
        logger.error(f"Error queuing LLM analysis: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Eroare internă: {str(e)}")


@router.get("/analyze-llm-status/{job_id}", response_model=Dict[str, Any])
async def get_analyze_llm_status(job_id: str):
    """
    Check the status of an LLM analysis job.
    Returns: {status: 'queued'|'processing'|'completed'|'failed'|'not_found', ...}
    """
    from ..logic.queue_manager import queue_manager
    import logging

    logger = logging.getLogger(__name__)

    try:
        status = queue_manager.get_job_status(job_id)
        logger.info(f"Job {job_id} status: {status.get('status')}")
        return status
    except Exception as e:
        logger.error(f"Error getting job status: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Eroare la verificarea statusului: {str(e)}")


def _generate_llm_data(session: Session, ultima: Any, candidate_count: int = None, custom_template: str = None):
    """
    Helper to generate the export data and optimized prompt.
    Args:
        candidate_count: Number of cases to include (default from settings)
        custom_template: Optional custom prompt template to use instead of settings
    """
    from sqlmodel import text
    import json
    import logging

    logger = logging.getLogger(__name__)

    # Get candidate count from settings if not provided
    if candidate_count is None:
        candidate_count = settings_manager.get_value('setari_llm', 'ai_filtering_llm_candidate_count', 5)

    # Limit to available IDs
    speta_ids_to_process = ultima.speta_ids[:candidate_count]

    # Construim query-ul doar pentru cele 5 ID-uri
    placeholders = ', '.join([f':id_{i}' for i in range(len(speta_ids_to_process))])
    query = text(f"SELECT id, obj FROM blocuri WHERE id IN ({placeholders})")

    params = {f'id_{i}': speta_id for i, speta_id in enumerate(speta_ids_to_process)}

    result = session.execute(query, params)
    results = result.mappings().all()

    spete_export = []
    spete_text_list = []

    for row in results:
        obj_data = row['obj']
        if isinstance(obj_data, str):
            try:
                obj = json.loads(obj_data)
            except json.JSONDecodeError:
                continue
        else:
            obj = obj_data

        situatia = (
            obj.get('text_situatia_de_fapt') or
            obj.get('situatia_de_fapt') or
            obj.get('situatie') or
            ""
        )

        # Siguranță: Tăiem textul dacă e prea lung pentru a nu bloca memoria
        if len(situatia) > 4000:
            situatia = situatia[:4000] + "... (text trunchiat)"

        text_individualizare = obj.get('text_individualizare', '')
        obiect = obj.get('obiect', '')

        speta_item = {
            'id': row['id'],
            'denumire': obj.get('denumire', f'Caz #{row["id"]}'),
            'situatia_de_fapt': situatia,
            'text_individualizare': text_individualizare,
            'obiect': obiect
        }
        spete_export.append(speta_item)

        spete_text_list.append(f"""
CAZ #{row['id']}:
OBIECT: {obiect}
SITUATIA DE FAPT: {situatia}
ELEMENTE DE INDIVIDUALIZARE: {text_individualizare}
--------------------------------------------------
""")

    prompt_spete = "\n".join(spete_text_list)

    # Get settings
    result_count = settings_manager.get_value('setari_llm', 'ai_filtering_result_count', 1)

    if custom_template:
        prompt_template = custom_template
    else:
        prompt_template = settings_manager.get_value(
            'setari_llm',
            'llm_prompt_template',
            # Default fallback value
            'Esti un judecator cu experienta, capabil sa analizeze spete juridice complexe si sa identifice precedente relevante.\\n\\nSARCINA TA:\\nAnalizeaza situatia de fapt prezentata de justitiabil mai jos si compar-o cu cele {num_candidates} spete pre-filtrate furnizate.\\nIdentifica cele mai relevante {num_results} spete asemanatoare cu situatia de fapt a utilizatorului, luand in considerare situatia de fapt, obiectul si elementele de individualizare.\\n\\nSITUATIA DE FAPT A JUSTITIABILULUI:\\n"{query_text}"\\n\\nLISTA DE SPETE PENTRU COMPARATIE ({num_candidates} Spete Pre-filtrate):\\n{prompt_spete}\\n\\nFORMATUL RASPUNSULUI:\\nGenereaza EXCLUSIV ID-urile spetelor selectate, separate prin virgula.\\nNu adauga niciun alt text, comentariu, explicatie sau introducere.\\nExemplu de raspuns valid pentru 1 rezultat: 123\\nExemplu de raspuns valid pentru 3 rezultate: 123, 456, 789'
        )

    # Format the template with actual data
    optimized_prompt = prompt_template.format(
        query_text=ultima.query_text,
        prompt_spete=prompt_spete,
        num_candidates=len(spete_export),
        num_results=result_count
    )

    return spete_export, optimized_prompt


@router.get("/materie-statistics", response_model=Dict[str, Any])
async def get_materie_statistics(
    session: Session = Depends(get_session),
    limit: int = 100
):
    """
    Get materie statistics showing which legal subjects are most frequently displayed.
    """
    from ..models import MaterieStatistics
    from sqlmodel import select
    import logging

    logger = logging.getLogger(__name__)

    try:
        # Query materie statistics, ordered by display_count descending
        statement = (
            select(MaterieStatistics)
            .order_by(MaterieStatistics.display_count.desc())
            .limit(limit)
        )

        results = session.exec(statement).all()

        # Format results
        statistics = [
            {
                'materie': stat.materie,
                'display_count': stat.display_count,
                'last_updated': stat.last_updated.isoformat() if stat.last_updated else None
            }
            for stat in results
        ]

        logger.info(f"Retrieved {len(statistics)} materie statistics")

        return {
            'success': True,
            'total_materii': len(statistics),
            'statistics': statistics
        }

    except Exception as e:
        logger.error(f"Error retrieving materie statistics: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Eroare la preluarea statisticilor: {str(e)}"
        )


@router.get("/materie-statistics/export")
async def export_materie_statistics(
    session: Session = Depends(get_session),
    limit: int = 1000
):
    """
    Export materie statistics to Excel file (.xlsx).
    """
    from ..models import MaterieStatistics
    from sqlmodel import select
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from datetime import datetime
    import logging

    logger = logging.getLogger(__name__)

    try:
        # Query materie statistics
        statement = (
            select(MaterieStatistics)
            .order_by(MaterieStatistics.display_count.desc())
            .limit(limit)
        )

        results = session.exec(statement).all()

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistici Materii"

        # Header styling
        header_fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")

        # Set headers
        headers = ["Nr.", "Obiect Spetă", "Număr Afișări", "Ultima Actualizare"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Add data
        for idx, stat in enumerate(results, 1):
            ws.cell(row=idx+1, column=1, value=idx)
            ws.cell(row=idx+1, column=2, value=stat.materie)
            ws.cell(row=idx+1, column=3, value=stat.display_count)
            ws.cell(row=idx+1, column=4, value=stat.last_updated.strftime("%Y-%m-%d %H:%M:%S") if stat.last_updated else "")

        # Adjust column widths
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 25

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        logger.info(f"Exported {len(results)} materie statistics to Excel")

        # Return as downloadable file
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f"attachment; filename=statistici_materii_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            }
        )

    except Exception as e:
        logger.error(f"Error exporting materie statistics: {e}", exc_info=True)
        raise HTTPException(
            status_code=500,
            detail=f"Eroare la exportul statisticilor: {str(e)}"
        )
